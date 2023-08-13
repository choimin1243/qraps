from fastapi import FastAPI, File, UploadFile, Form
from fastapi.responses import StreamingResponse
from io import BytesIO
import openpyxl
from urllib.parse import quote
from fastapi.responses import HTMLResponse
import pandas as pd
import re
import os
from openpyxl.utils.dataframe import dataframe_to_rows

app = FastAPI()


@app.get("/")
async def read_root():
    return HTMLResponse(content=open("index.html", "r", encoding="utf-8").read())

@app.post("/upload/")
async def upload_file(file: UploadFile = File(...), text_data: str = Form(...),additional_text: str=Form(...)):
    try:
        print(UploadFile)
        print(text_data)
        print(additional_text)
        uploaded_file = file.file.read()

        # 엑셀 파일 열기
        wb = openpyxl.load_workbook(BytesIO(uploaded_file))
        sheet = wb.active

        last_column = sheet.max_column
        last_row = sheet.max_row

        data = []
        location_column_index = []  # "Location No."가 위치하는 컬럼의 인덱스를 저장할 변수
        part_number = []

        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=last_row, max_col=last_column):
            # 한 행의 데이터를 저장할 리스트를 생성합니다.
            row_data = []
            for idx, cell in enumerate(row, 1):
                if cell.value == additional_text:
                    location_column_index.append(idx)

                row_data.append([cell.value])

            data.append(row_data)

        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=last_row, max_col=last_column):
            # 한 행의 데이터를 저장할 리스트를 생성합니다.
            for idx, cell in enumerate(row, 1):
                if cell.value is not None and isinstance(cell.value, str):
                    cell.value = cell.value.lower()
                    if cell.value == "package":
                        part_number = idx
                        print(idx, "@@")

        if part_number == []:
            part_number = 15

        print(part_number)

        def remove_duplicates(input_list):
            return list(set(input_list))

        patternwat = r"(\d+(?:/\d+)?(?:\.\d+)?)\s*(w|kw|mw)"
        patternnp = r"(?<!\w)(?<!\d)\d+(?:\.\d+)?(?:\s*(?:pF|nF|uF|µF|UF|p|n|u|µ))(?!\w)"
        patternv = r"(\d+(?:\.\d+)?)\s*(?:[kK]?[mM]?[vV])"

        tolerance_values = ["J", "F", "A", "B", "G", "M", "Z"]

        # 수정된 정규식
        pattern_tor = r"(?<![A-Za-z0-9.,-])(?:{})(?![A-Za-z0-9.,])".format("|".join(tolerance_values))

        patternAed = r"([-+]?\d+(?:\.\d+)?)\s*([mµ]?[AaKk])"
        patterntemp = r'\d+(?:\.\d+)?\s*℃'

        location_column_index = remove_duplicates(location_column_index)
        pattern_kv = r"kv"
        pattern_v = r"\d+(?:\.\d+)?(?=\s*(?i)v)"

        result_data = []

        character = text_data

        pattern = r"(?<!\S)" + character + "(\d+)"

        list_row = []

        for i in range(len(data)):
            try:
                datas = data[i][location_column_index[0] - 1]
                if datas[0] != None:
                    parsed_data = [row.replace(" ", "").split(",") for row in datas[0].split("\n")]
                    flattened_data = [item for sublist in parsed_data for item in sublist]
                    if re.findall(pattern, flattened_data[0], re.IGNORECASE):
                        list_row.append(i)
                        for s in range(len(flattened_data)):
                            if flattened_data[s] != '':
                                result_data.append([i, flattened_data[s].strip()])


            except:
                pass

        voltage_number = 1
        wat_number = 1
        resistance_number = 1
        tolerance_number = 1
        nlp = 1
        part_num = 1
        list_table_number = ["No", "REF NO"]

        pattern_kv = r"kv"
        pattern_v = r"(\d+(?:\.\d+)?)\s*(?:[vV])"
        patternv = r"(\d+(?:\.\d+)?)\s*(?:[kK]?[mM]?[vV])"
        print(list_row, "@@@")

        for i in range(len(list_row)):
            data_item = data[list_row[i]]
            print(list_row, "@@@")

            for s in range(len(data_item)):
                try:
                    something = re.search(patternv, data_item[s][0])
                    voltage_value = something.group(0)
                    for k in range(len(result_data)):
                        if result_data[k][0] == list_row[i]:
                            voltage_value = something.group(0)
                            matches = re.search(pattern_kv, voltage_value, re.IGNORECASE)
                            match = re.search(pattern_v, voltage_value, re.IGNORECASE)
                            if match:
                                voltage_number = 2
                                matches_data = re.findall(pattern_v, voltage_value)
                                result_data[k].append(int(matches_data[0]))
                            if matches:
                                voltage_number = 2
                                matches_num = re.findall(patternv, voltage_value, re.IGNORECASE)
                                result_data[k].append(int(matches_num[0]) * 1000)

                    break

                except:
                    pass

        if voltage_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")
            list_table_number.append("VOLTAGE")

        print(result_data)

        for i in range(len(list_row)):
            data_item = data[list_row[i]]

            for s in range(len(data_item)):
                try:
                    matches = re.findall(patternwat, data_item[s][0], re.IGNORECASE)
                    if matches:
                        wat_number = 2

                        combined_values = [f"{match[0]}{match[1].lower()}" for match in matches]
                        combined_result = " ".join(combined_values)
                        wat = combined_result
                        for k in range(len(result_data)):
                            if result_data[k][0] == list_row[i]:
                                result_data[k].append(wat)
                        break
                    else:
                        wat = ""

                except:
                    wat = ""

        if wat_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")
            list_table_number.append("RATED_POWER[W]")

        for i in range(len(list_row)):
            data_item = data[list_row[i]]
            print(data_item)
            for s in range(len(data_item)):
                try:
                    tolerance_value = ""
                    match = re.search(pattern_tor, data_item[s][0])
                    if match:
                        tolerance_number = 2
                        tolerance_value = match.group(0)
                        print(tolerance_value)
                        for k in range(len(result_data)):
                            if result_data[k][0] == list_row[i]:
                                result_data[k].append(tolerance_value)
                        break
                    else:
                        resistance_value = ""


                except:
                    resistance_value = ""

        if tolerance_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")
            list_table_number.append("TOLERANCE")

        patternom = r"(?:,\s*)?(\d+(?:\.\d+)?)(?:\s*(?:㏀|Ω|k㏀|kΩ|mΩ|㏁))\s*\*?\d?"

        for i in range(len(list_row)):
            data_item = data[list_row[i]]
            print(data_item)
            for s in range(len(data_item)):
                try:
                    matchnorm = re.search(patternom, data_item[s][0])

                    if matchnorm:
                        resistance_number = 2
                        print(matchnorm)
                        resistance_value = matchnorm.group(0)
                        for k in range(len(result_data)):
                            if result_data[k][0] == list_row[i]:
                                result_data[k].append(resistance_value)
                        break
                    else:
                        resistance_value = ""
                        print(resistance_value)


                except:
                    resistance_value = ""

        if resistance_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")
            list_table_number.append("RESISTANCE")

        print(list_table_number)

        temp_number = 1

        for i in range(len(list_row)):
            data_item = data[list_row[i]]
            print("@@", data_item)
            for s in range(len(data_item)):
                try:
                    tolerance_value = ""
                    match = re.search(patterntemp, data_item[s][0], re.IGNORECASE)
                    if match:
                        temp_number = 2
                        tmp_value = match.group(0)
                        for k in range(len(result_data)):
                            if result_data[k][0] == list_row[i]:
                                result_data[k].append(tmp_value)
                        break
                    else:
                        resistance_value = ""


                except:
                    resistance_value = ""

        if temp_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")

            list_table_number.append("TEMPERATURE")

        #
        nlp_number = 1
        for i in range(len(list_row)):
            data_item = data[list_row[i]]
            print("@@", data_item)
            for s in range(len(data_item)):
                try:
                    tolerance_value = ""
                    match = re.search(patternnp, data_item[s][0], re.IGNORECASE)
                    if match:
                        nlp_number = 2
                        nlp_value = match.group(0)
                        print(tolerance_value)
                        for k in range(len(result_data)):
                            if result_data[k][0] == list_row[i]:
                                result_data[k].append(nlp_value)
                        break
                    else:
                        resistance_value = ""


                except:
                    resistance_value = ""

        if nlp_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")

            list_table_number.append("CAPACITANCE")
        #
        #
        #
        #
        #
        #
        #
        for i in range(len(list_row)):
            data_item = data[list_row[i]]
            for k in range(len(result_data)):
                if result_data[k][0] == list_row[i]:
                    result_data[k].append(data_item[part_number - 1][0])

        list_table_number.append("PACKAGE")

        pattern_caps = r"(X7R|X5R|COG|NPO|X5S|X6S)"

        Grade = 1
        for i in range(len(list_row)):
            data_item = data[list_row[i]]
            for s in range(len(data_item)):
                try:
                    tolerance_value = ""
                    match = re.search(pattern_caps, data_item[s][0], re.IGNORECASE)
                    if match:
                        Grade = 2
                        Grade_value = match.group(0)
                        for k in range(len(result_data)):
                            if result_data[k][0] == list_row[i]:
                                result_data[k].append(Grade_value)
                        break
                    else:
                        resistance_value = ""


                except:
                    resistance_value = ""

        if Grade == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")

            list_table_number.append("GRADE")

        for row in result_data:
            if row[1].isdigit():
                row[1] = character + row[1]

        print(list_table_number)

        for row in result_data:
            num = len(character)
            row[0] = int(row[1][num:])

        result_data.insert(0, list_table_number)

        print(result_data)

        df = pd.DataFrame(result_data[1:], columns=result_data[0])

        # 첫 번째 열을 기준으로 정렬
        sorted_df = df.sort_values(by='No')

        A_table = ["No", "REF NO", "PACKAGE", "RATED_POWER[W]", "TOLERANCE", "RESISTANCE"]
        B_table = ["No", "REF NO", "PACKAGE", "CAPACITANCE", "VOLTAGE", "GRADE", "TOLERANCE", "TEMPERATURE"]

        if character == "R":
            type = A_table
        else:
            type = B_table

        df = pd.DataFrame(sorted_df, columns=type)
        print(df)




        new_sheet = wb.create_sheet(title="New Sheet")
        for row in dataframe_to_rows(df, index=False, header=True):
            new_sheet.append(row)

        # 업데이트된 파일을 메모리에 저장
        updated_file_content = BytesIO()
        wb.save(updated_file_content)

        # 파일 이름 URL 인코딩
        encoded_file_name = quote(file.filename, safe="")

        # Generate a StreamingResponse for downloading the file
        response = StreamingResponse(
            iter([updated_file_content.getvalue()]),
            headers={
                "Content-Disposition": f"attachment; filename={encoded_file_name}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )


        # Delete the created Excel file

        return response


    except Exception as e:
        return {"error": str(e)}

